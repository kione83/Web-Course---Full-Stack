"""Set imports"""
import requests
import json
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import time

"""Declare Variables"""
query_end = ""
wb = load_workbook('Employment Tracker.xlsx') #active workbook
sheets =['WD DK Jul','WD DK May']
for sheet in sheets:
	wb.active = wb[sheet]  # active worksheet
ws = wb.active
url = "https://linkedin-profile-data.p.rapidapi.com/linkedin-data"
querystring = {"url":"https://www." + query_end} #querystring for API call
headers = {
	"X-RapidAPI-Key": "13860b3cfcmsh53f92e505445da2p14f18ejsn56b146dbccdb",
	"X-RapidAPI-Host": "linkedin-profile-data.p.rapidapi.com"
} #headers for API call




#loop through d column and make API call, then update the h column with the response
def LinkedIn_Call():
		query_end = cell.value
		querystring = {"url": "https://www." + query_end}
		response = requests.request("GET", url, headers=headers, params=querystring)
		print(response.url)	
		json_response = json.loads(response.text)
		print(json_response)
		if json_response != "Server is busy. Please try again later.":
			year_started_job = json_response["experiences"][0]["starts_at"]["year"]
			# write this to the excel file in the L through O columns
			ws.cell(row=cell.row, column=11).value = year_started_job
			ws.cell(row=cell.row, column=12).value= json_response["experiences"][0]["starts_at"]["month"]
			print(query_end)
			#wait 15 seconds before making the next API call
			time.sleep(10)
			#save the excel file
			wb.save('Employment Tracker.xlsx')
		else:
			print(response.status_code)

def stripCells():
	#strip out the http://www. from the cell, the https://www. from the cell, and the www. from the cell .. 
	# all cells in the D column should start with linkedin.com/in/ and end with a username or profile id
	for cell in ws['D']:
		if cell.value is not None:
			if "http://www." in cell.value:
				cell.value = cell.value.replace("http://www.", "")	
			elif "https://www." in cell.value:
				cell.value = cell.value.replace("https://www.", "")
			elif "www." in cell.value:
				cell.value = cell.value.replace("www.", "")
			else:
				pass


for sheet in sheets:
#iterate the rows of column D starting with 2
	for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
		#check for empty cells
		if row[0].value == "None" or row[0].value == "n/a":
			pass
		else:	
			for cell in row:		
				LinkedIn_Call()





# sheet_names = ['WD DK-1 Feb','WD DK-2 Feb','Cybersecurity PANW Roster Mar 7','Web Dev Boeing Mar 7','Web Dev AT&T Mar 14','Cybersecurity PANW Roster Apr 4','Security+ Bootcamp Apr 04','WD DK May','Cybersecurity PANW Roster Jun 13','Security+ Bootcamp Jun 13','WD Boeing June','WD DK Jul','Security+ Bootcamp Aug 1','WD DK Aug','Cybersecurity PANW Roster Aug 15','PANW Roster Oct 3','WD DK Oct','ITP DK May','ITP DK Sept']

#main
if __name__ == "__main__":
	stripCells()
	LinkedIn_Call()
	wb.save('Employment Tracker.xlsx')










